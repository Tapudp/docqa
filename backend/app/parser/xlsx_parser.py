import io

from app.parser.base import DocumentParser, ParseResult

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class XlsxParser(DocumentParser):
    def supports(self, mime_type: str) -> bool:
        return mime_type == _XLSX_MIME

    async def parse(self, data: bytes, mime_type: str) -> ParseResult:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        pages: list[str] = []

        for sheet in wb.worksheets:
            lines: list[str] = [f"Sheet: {sheet.title}"]
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c for c in cells):
                    lines.append("\t".join(cells))
            pages.append("\n".join(lines))

        wb.close()

        if not pages:
            pages = [""]

        return ParseResult(total_pages=len(pages), page_texts=pages, failed_pages=[])
